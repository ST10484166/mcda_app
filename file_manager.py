"""File I/O for projects: JSON save/load, autosave, Excel/PDF export."""

from __future__ import annotations
import json
import os
from datetime import datetime
from typing import List, Dict, Any, Optional

from models import Scenario

AUTOSAVE_FILENAME = "autosave_mcda.json"


class ProjectFile:
    """A project = a named collection of Scenarios, saved as one JSON file."""

    def __init__(self, scenarios: Optional[List[Scenario]] = None,
                 active_scenario_name: Optional[str] = None):
        self.scenarios: List[Scenario] = scenarios or []
        self.active_scenario_name = active_scenario_name or (
            self.scenarios[0].name if self.scenarios else None
        )

    def get_active(self) -> Optional[Scenario]:
        for s in self.scenarios:
            if s.name == self.active_scenario_name:
                return s
        return self.scenarios[0] if self.scenarios else None

    def add_scenario(self, scenario: Scenario) -> None:
        """Add or overwrite a scenario and make it active."""
        self.scenarios = [s for s in self.scenarios if s.name != scenario.name]
        self.scenarios.append(scenario)
        self.active_scenario_name = scenario.name

    def delete_scenario(self, name: str) -> None:
        self.scenarios = [s for s in self.scenarios if s.name != name]
        if self.active_scenario_name == name:
            self.active_scenario_name = self.scenarios[0].name if self.scenarios else None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "active_scenario_name": self.active_scenario_name,
            "scenarios": [s.to_dict() for s in self.scenarios],
        }

    @staticmethod
    def from_dict(d: Dict[str, Any]) -> "ProjectFile":
        scenarios = [Scenario.from_dict(sd) for sd in d.get("scenarios", [])]
        return ProjectFile(scenarios=scenarios,
                            active_scenario_name=d.get("active_scenario_name"))

    def save(self, path: str) -> None:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.to_dict(), f, indent=2, ensure_ascii=False)

    @staticmethod
    def load(path: str) -> "ProjectFile":
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return ProjectFile.from_dict(data)

    def autosave(self, directory: str) -> str:
        os.makedirs(directory, exist_ok=True)
        path = os.path.join(directory, AUTOSAVE_FILENAME)
        self.save(path)
        return path


# -------------------------------------------------------------------------
# Excel export
# -------------------------------------------------------------------------

def export_to_excel(scenario: Scenario, result, path: str) -> None:
    """Export the criteria matrix + summary results to an .xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Criteria Matrix"

    headers = ["Criterion", "Weight (%)", f"{scenario.option_a_label} Score",
               f"{scenario.option_b_label} Score", "Weighted A", "Weighted B",
               "Notes A", "Notes B"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A5E2E", end_color="4A5E2E", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True)

    for c in scenario.criteria:
        ws.append([c.name, c.weight, c.score_a, c.score_b,
                   round(c.weighted_a(), 3), round(c.weighted_b(), 3),
                   c.notes_a, c.notes_b])

    widths = [28, 12, 16, 16, 12, 12, 40, 40]
    for col, width in zip("ABCDEFGH", widths):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.append(["Option A total", round(result.total_a, 3)])
    ws2.append(["Option B total", round(result.total_b, 3)])
    ws2.append(["Difference (A - B)", round(result.difference, 3)])
    ws2.append(["Percentage difference", f"{result.pct_difference:.2f}%"])
    ws2.append(["Recommended option", result.recommended or "Tie / none"])
    ws2.append(["Confidence", result.confidence_label])
    ws2.append(["Option A feasible", result.feasibility.option_a_feasible])
    ws2.append(["Option B feasible", result.feasibility.option_b_feasible])
    if result.feasibility.failed_a:
        ws2.append(["Option A failed thresholds", ", ".join(result.feasibility.failed_a)])
    if result.feasibility.failed_b:
        ws2.append(["Option B failed thresholds", ", ".join(result.feasibility.failed_b)])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 45

    ws3 = wb.create_sheet("Critical Thresholds")
    ws3.append(["Threshold", "Option A Pass?", "Option B Pass?", "Notes"])
    for t in scenario.thresholds:
        ws3.append([t.name, "Yes" if t.pass_a else "No", "Yes" if t.pass_b else "No", t.notes])
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["D"].width = 45

    wb.save(path)


# -------------------------------------------------------------------------
# PDF export
# -------------------------------------------------------------------------

def export_to_pdf(scenario: Scenario, result, chart_image_paths: List[str], path: str) -> None:
    """Export a report PDF: text summary + criteria table + embedded chart
    images (one per page)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader

    c = canvas.Canvas(path, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 16)
    c.drawString(2 * cm, height - 2 * cm, "MCDA Decision Report")
    c.setFont("Helvetica", 10)
    c.drawString(2 * cm, height - 2.7 * cm,
                 f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    y = height - 3.6 * cm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2 * cm, y, f"Scenario: {scenario.name}")
    y -= 0.8 * cm
    c.setFont("Helvetica", 11)
    lines = [
        f"{scenario.option_a_label}: {result.total_a:.2f} / 10",
        f"{scenario.option_b_label}: {result.total_b:.2f} / 10",
        f"Difference: {result.difference:+.2f}  ({result.pct_difference:.1f}%)",
        f"Recommended: {result.recommended or 'Tie / none'}",
        f"Confidence: {result.confidence_label}",
    ]
    for line in lines:
        c.drawString(2 * cm, y, line)
        y -= 0.6 * cm

    if not result.feasibility.option_a_feasible or not result.feasibility.option_b_feasible:
        y -= 0.3 * cm
        c.setFont("Helvetica-Bold", 11)
        c.drawString(2 * cm, y, "Feasibility Gate Warnings:")
        y -= 0.6 * cm
        c.setFont("Helvetica", 10)
        if result.feasibility.failed_a:
            c.drawString(2 * cm, y, f"Option A blocked by: {', '.join(result.feasibility.failed_a)}")
            y -= 0.5 * cm
        if result.feasibility.failed_b:
            c.drawString(2 * cm, y, f"Option B blocked by: {', '.join(result.feasibility.failed_b)}")
            y -= 0.5 * cm

    y -= 0.4 * cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2 * cm, y, "Criteria")
    y -= 0.5 * cm
    c.setFont("Helvetica", 9)
    for cr in scenario.criteria:
        if y < 3 * cm:
            c.showPage()
            y = height - 2 * cm
            c.setFont("Helvetica", 9)
        c.drawString(
            2 * cm, y,
            f"{cr.name}: weight {cr.weight:.1f}%, A={cr.score_a}, B={cr.score_b}, "
            f"wA={cr.weighted_a():.2f}, wB={cr.weighted_b():.2f}",
        )
        y -= 0.45 * cm

    for img_path in chart_image_paths:
        if os.path.exists(img_path):
            c.showPage()
            img = ImageReader(img_path)
            iw, ih = img.getSize()
            scale = min((width - 4 * cm) / iw, (height - 4 * cm) / ih)
            c.drawImage(img, 2 * cm, height - 2 * cm - ih * scale,
                        width=iw * scale, height=ih * scale)

    c.save()
